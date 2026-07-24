import os
import openpyxl

def get_ndk_apis_from_excel(excel_file='data/ndk_api.xlsx', class_column= 3, api_column= 0):
    """
    description: get apis dict from single excel file
    """

    ret = {}
    apis_dict = {}
    file_name = os.path.split(excel_file)[-1].split('.')[0]
    try:
        wb = openpyxl.load_workbook(excel_file)
    except:
        print(f'load ndk file {excel_file} failed')
        return ret

    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        all_data = tuple(sheet.rows)
        datas = all_data[1:]
        max_columns = sheet.max_column
        pros = all_data[0]
        for data in datas:
            if data[class_column].value:
                if '::' in data[class_column].value:
                    api_name = data[class_column].value.split('::')[-1]
                elif '.' in data[class_column].value:
                    temp_list = data[class_column].value.split('.')
                    api_name = temp_list[0] + '.' + temp_list[1]
                else:
                    api_name = data[api_column].value
            else:
                api_name = data[api_column].value
            api_info = {}
            for column in range(max_columns):
                property_data = data[column].value
                property_name = pros[column].value
                api_info[property_name] = property_data
            if 'SUPPORTED' in api_info['status']:
                print(api_info['symbol'], ' : ', api_info['status'])
                continue
            ret[api_name] = api_info
            apis_dict[file_name] = ret
    print(f'loading excel {excel_file} data finished')
    # print(ret)
    return apis_dict

if __name__ == '__main__':
    # ndk_apis = get_ndk_apis_from_excel()
    # print(ndk_apis.keys())
    # print(ndk_apis)
    libs_apis = get_ndk_apis_from_excel('./black_list/ndk.xlsx')
    print(libs_apis.keys())
    print(libs_apis)
